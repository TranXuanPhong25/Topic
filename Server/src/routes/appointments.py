from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
import os
import uuid
import openpyxl
from datetime import datetime

router = APIRouter(prefix="/appointments", tags=["Appointments"])

DATA_FOLDER = "data"
FILE_PATH = os.path.join(DATA_FOLDER, "appointments.xlsx")


# Initialize Excel storage
def init_excel():
    if not os.path.exists(DATA_FOLDER):
        os.makedirs(DATA_FOLDER)

    if not os.path.exists(FILE_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"
        ws.append(["ID", "PatientName", "Reason", "Time", "Phone"])
        wb.save(FILE_PATH)
    else:
        # Nếu file đã có nhưng chưa có cột "Phone" → tự thêm
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if "Phone" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="Phone")
            wb.save(FILE_PATH)


# Generate UUID4
def generate_uuid():
    return str(uuid.uuid4())


# Pydantic Model
class AppointmentRequest(BaseModel):
    patient_name: str
    reason: str
    time: str      # "2025-12-01 10:30"
    phone: str     # số điện thoại


# Check duplicate time
def is_duplicate_time(time_str, ignore_id=None):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_id, _, _, existing_time, _ = row
        if existing_time == time_str and existing_id != ignore_id:
            return True
    return False


# ➤ CREATE APPOINTMENT
@router.post("/create")
async def create_appointment(request: AppointmentRequest):
    init_excel()

    # Check duplicate time
    if is_duplicate_time(request.time):
        raise HTTPException(
            status_code=400,
            detail=f"Đã có lịch khám lúc {request.time}. Vui lòng chọn thời gian khác."
        )

    appointment_id = generate_uuid()

    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    ws.append([
        appointment_id,
        request.patient_name,
        request.reason,
        request.time,
        request.phone
    ])
    wb.save(FILE_PATH)

    return {
        "id": appointment_id,
        "patient_name": request.patient_name,
        "reason": request.reason,
        "time": request.time,
        "phone": request.phone
    }


# ➤ LIST ALL APPOINTMENTS
@router.get("/list")
async def list_appointments():
    init_excel()
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    appointments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        appointments.append({
            "id": row[0],
            "patient_name": row[1],
            "reason": row[2],
            "time": row[3],
            "phone": row[4]
        })

    return appointments


# ➤ GET BY ID
@router.get("/{appointment_id}")
async def get_appointment(appointment_id: str):
    init_excel()
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == appointment_id:
            return {
                "id": row[0],
                "patient_name": row[1],
                "reason": row[2],
                "time": row[3],
                "phone": row[4]
            }

    raise HTTPException(status_code=404, detail="Không tìm thấy lịch khám")


# ➤ UPDATE APPOINTMENT
@router.put("/{appointment_id}")
async def update_appointment(appointment_id: str, request: AppointmentRequest):
    init_excel()

    # Kiểm tra trùng giờ (trừ chính ID này)
    if is_duplicate_time(request.time, ignore_id=appointment_id):
        raise HTTPException(
            status_code=400,
            detail=f"Đã có lịch khám lúc {request.time}. Vui lòng chọn thời gian khác."
        )

    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == appointment_id:
            row[1].value = request.patient_name
            row[2].value = request.reason
            row[3].value = request.time
            row[4].value = request.phone
            wb.save(FILE_PATH)

            return {
                "id": appointment_id,
                "patient_name": request.patient_name,
                "reason": request.reason,
                "time": request.time,
                "phone": request.phone
            }

    raise HTTPException(status_code=404, detail="Không tìm thấy lịch khám")


# ➤ DELETE APPOINTMENT
@router.delete("/{appointment_id}")
async def delete_appointment(appointment_id: str):
    init_excel()
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == appointment_id:
            ws.delete_rows(row)
            wb.save(FILE_PATH)
            return {"message": "Đã xóa lịch khám thành công"}

    raise HTTPException(status_code=404, detail="Không tìm thấy lịch khám")
